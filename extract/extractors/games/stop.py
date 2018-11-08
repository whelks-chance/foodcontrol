import statistics
from collections import defaultdict
from openpyxl import Workbook
from pprint import pprint

from .gamedataextractor import GameDataExtractor
from .session_event_log import SessionEventLog
from keypath_extractor import Keypath
from spreadsheet import Spreadsheet
from utils import irange


class AbstractStopDataExtractor(GameDataExtractor):
    """The abstract base class for all STOP games"""

    @staticmethod
    def get_value_keypaths():
        return [
            Keypath('data.captureDate', 'Capture Date'),
        ]

    session_event_log = SessionEventLog()
    session_duration = 0
    durations = defaultdict(list)
    raw_count = {
        'on': defaultdict(int),
        'off': defaultdict(int)
    }

    @staticmethod
    def numericify(n):
        if n is None:
            n = 0
        return n

    @staticmethod
    def denominator(n):
        if n:
            return n
        else:
            return 1

    @staticmethod
    def mean(numbers):
        if numbers and len(numbers) > 0:
            return statistics.mean(numbers)
        else:
            return 0

    @staticmethod
    def remove_none_values(values):
        return [d for d in values if d is not None]

    def clear(self):
        self.session_duration = 0
        self.durations.clear()
        self.session_trial_type_counts.clear()
        self.trial_type_block_count.clear()
        self.block_item_type_counts.clear()
        self.raw_count['on'].clear()
        self.raw_count['off'].clear()

    def check(self, row):
        super(AbstractStopDataExtractor, self).check(row)

        def check_trials_count():
            number_of_rounds = 4
            old_number_of_trials = 48
            new_number_of_trials = 24
            session_events = self.get_keypath_value(row, 'data.0.sessionEvents')
            assert len(session_events) == (number_of_rounds * old_number_of_trials) or (number_of_rounds * new_number_of_trials)

        check_trials_count()

    def check(self, row):
        self.check_tap_responses(row)

    def get_session_events(self, row):
        """
        Handle inconsistently formatted data structures:
        data -> object vs data -> [ object ]
        """
        session_events = None
        try:
            session_events = self.get_keypath_value(row, 'data.0.sessionEvents')
        except KeyError:
            try:
                session_events = self.get_keypath_value(row, 'data.sessionEvents')
            except KeyError:
                pass
        return session_events

    @staticmethod
    def within_stimulus_boundary(session_event, item_radius=95, prefix='tapResponsePosition'):
        tx = float(session_event['{}X'.format(prefix)])
        ty = float(session_event['{}Y'.format(prefix)])
        ix = session_event['itemPositionX']
        iy = session_event['itemPositionY']
        # print(tx, ty, ix, iy)
        return ((tx - ix) ** 2) + ((ty - iy) ** 2) < (item_radius ** 2)

    def outside_stimulus_boundary(self, session_event, prefix='tapResponsePosition'):
        return not self.within_stimulus_boundary(session_event, prefix=prefix)

    def check_tap_responses(self, row):
        # trs = tap response start
        checks = {
            'GO': {
                'CORRECT_GO': lambda trs, session_event: trs > 0 and self.within_stimulus_boundary(session_event),
                'INCORRECT_GO': lambda trs, session_event: trs == 0,
                'MISS_GO': lambda trs, session_event: trs > 0 and self.outside_stimulus_boundary(session_event),
            },
            'STOP': {
                'CORRECT_STOP': lambda trs, session_event: trs == 0,
                'INCORRECT_STOP': lambda trs, session_event: trs > 0 and self.within_stimulus_boundary(session_event),
                'MISS_STOP': lambda trs, session_event: trs > 0 and self.outside_stimulus_boundary(session_event),
            }
        }
        session_events = self.get_session_events(row)
        for session_event in session_events:
            trial_type = session_event['trialType']
            tap_response_type = session_event['tapResponseType']
            tap_response_start = self.numericify(session_event['tapResponseStart'])
            check_result = checks[trial_type][tap_response_type](tap_response_start, session_event)
            self.session_event_log.log_if_check_failed(check_result, session_event)

    def check_points(self, row):
        points = {
            'GO': {
                'CORRECT_GO': 20,
                'INCORRECT_GO': -50,
                'MISS_GO': -20,
            },
            'STOP': {
                'CORRECT_STOP': 50,
                'INCORRECT_STOP': -50,
                'MISS_STOP': -50,
            }
        }
        running_total = 0
        session_events = self.get_session_events(row)
        for session_event in session_events:
            trial_type = session_event['trialType']
            tap_response_type = session_event['tapResponseType']
            points_this_trial = session_event['pointsThisTrial']
            assert points_this_trial == points[trial_type][tap_response_type]
            running_total += points_this_trial
            points_running_total = session_event['pointsRunningTotal']
            assert points_running_total == running_total

    def calculate_ssrt(self, row):
        session_events = self.get_session_events(row)
        if session_events:

            # Mean SSRT
            tap_response_start_total = 0
            stop_signal_delay_total = 0
            stop_signal_onset_total = 0
            for session_event in session_events:
                trial_type = session_event['trialType']
                tap_response_start = self.numericify(session_event['tapResponseStart'])
                if trial_type == 'GO' and tap_response_start > 0:
                    tap_response_start_total += tap_response_start
                stop_signal_delay = self.numericify(session_event['stopSignalDelay'])
                stop_signal_delay_total += stop_signal_delay
                stop_signal_onset = self.numericify(session_event['stopSignalOnset'])
                stop_signal_onset_total += stop_signal_onset
            mean_tap_response_start = tap_response_start_total / len(session_events)
            mean_stop_signal_delay = stop_signal_delay_total / len(session_events)
            mean_stop_signal_onset = stop_signal_onset_total / len(session_events)
            mean_ideal_ssrt = mean_tap_response_start - mean_stop_signal_delay  # What the SSRT should be
            mean_actual_ssrt = mean_tap_response_start - mean_stop_signal_onset  # What the SSRT actually is
            print('\n IDEAL MEAN SSRT:', mean_ideal_ssrt)
            print('ACTUAL MEAN SSRT:', mean_actual_ssrt)

            # Integration SSRT
            incorrect_stop_trials_count = 0
            for session_event in session_events:
                tap_response_type = session_event['tapResponseType']
                if tap_response_type == 'INCORRECT_STOP' or tap_response_type == 'MISS_STOP':
                    incorrect_stop_trials_count += 1
            print('INCORRECT STOP TRIALS:', incorrect_stop_trials_count)

    def calculate(self, row):
        super(AbstractStopDataExtractor, self).calculate(row)

        def calculate_durations():
            # "Durations"

            def calculate_session_duration():
                session_start = self.get_keypath_value(row, 'data.0.sessionStart')
                session_end = self.get_keypath_value(row, 'data.0.sessionEnd')
                self.session_duration = session_start - session_end

            def record_duration(key, value):
                self.durations[key].append(value)

            def calculate_trial_durations():

                def not_none(a, b):
                    return a is not None and b is not None

                session_events = self.get_session_events(row)
                if session_events:
                    previous_session_event = None
                    for session_event in session_events:
                        trial_start = session_event['trialStart']
                        trial_end = session_event['trialEnd']
                        trial_duration = trial_end - trial_start
                        record_duration('trial', trial_duration)

                        stop_signal_onset = session_event['stopSignalOnset']
                        stop_signal_offset = session_event['stopSignalOffset']
                        if not_none(stop_signal_onset, stop_signal_offset):
                            stop_signal_duration = stop_signal_offset - stop_signal_onset
                            record_duration('stop_signal', stop_signal_duration)

                        stimulus_onset = session_event['stimulusOnset']
                        stimulus_offset = session_event['stimulusOffset']
                        if not_none(stimulus_onset, stimulus_offset):
                            stimulus_duration = stimulus_offset - stimulus_onset
                            record_duration('stimulus', stimulus_duration)

                        # Difference between signal onset and stop signal delay
                        stop_signal_delay = session_event['stopSignalDelay']
                        if not_none(stop_signal_onset, stop_signal_delay):
                            signal_stop_difference = stop_signal_onset - stop_signal_delay
                            record_duration('signal_stop_difference', signal_stop_difference)

                        # Duration between signal offset and stimulus offset
                        if not_none(stimulus_offset, stop_signal_offset):
                            stimulus_stop_difference = stimulus_offset - stop_signal_offset
                            record_duration('stimulus_stop_difference', stimulus_stop_difference)

                        if previous_session_event:
                            previous_trial_start = previous_session_event['trialStart']
                            inter_trial_duration = trial_start - previous_trial_start
                            record_duration('inter_trial', inter_trial_duration)

                        previous_session_event = session_event

            def calculate_trial_duration_stats():

                def calculate_stats(durations_key):
                    durations = self.remove_none_values(self.durations[durations_key])
                    return {
                        'min': min(durations),
                        'max': max(durations),
                        'mean': statistics.mean(durations),
                        'stdev': statistics.stdev(durations),
                    }

                trial_categories = [
                    'trial',
                    'stop_signal',
                    'stimulus',
                    'signal_stop_difference',
                    'stimulus_stop_difference',
                    'inter_trial'
                ]

                self.trial_stats = {}
                for trial_category in trial_categories:
                    self.trial_stats[trial_category] = calculate_stats(trial_category)

            calculate_session_duration()
            calculate_trial_durations()
            calculate_trial_duration_stats()

        def count_trial_and_types():
            self.trial_count = 0
            self.raw_round_trial_counts = defaultdict(set)
            self.block_trial_type_counts = defaultdict(lambda: defaultdict(int))  # dict of int dict
            self.block_item_type_counts = defaultdict(lambda: defaultdict(int))   # dict of int dict

            session_events = self.get_keypath_value(row, 'data.0.sessionEvents')
            for session_event in session_events:
                self.trial_count += 1

                # Create a set of the trial IDs so we can count the elements
                block_id = session_event['roundID']
                trial_id = session_event['trialID']
                self.raw_round_trial_counts[block_id].add(trial_id)

                # Record the block-level trial type counts (GO/STOP)
                trial_type = session_event['trialType']
                self.block_trial_type_counts[block_id][trial_type] += 1

                # Record the block-level item type counts (HEALTHY/NON-HEALTHY)
                item_type = session_event['itemType']
                selected = session_event['selected']
                if item_type == 'HEALTHY':
                    self.block_item_type_counts[block_id]['HEALTHY'] += 1
                    if selected == 'random':
                        self.block_item_type_counts[block_id]['HEALTHY_RANDOM'] += 1
                    else:
                        self.block_item_type_counts[block_id]['HEALTHY_NOT_RANDOM'] += 1
                if item_type == 'NON_HEALTHY':
                    self.block_item_type_counts[block_id]['NON_HEALTHY'] += 1

            # Calculate the block-level trial type percentages
            self.block_trial_type_percentages = defaultdict(lambda: defaultdict(float))  # dict of float dict
            for block_id_key, items in self.block_trial_type_counts.items():
                block_total = 0
                for item_key, item_count in items.items():
                    block_total += item_count
                for item_key, item_count in items.items():
                    self.block_trial_type_percentages[block_id_key][item_key] = self.block_trial_type_counts[block_id_key][item_key] / block_total

            # Calculate the block-level item type percentages
            self.block_item_type_percentages = defaultdict(lambda: defaultdict(float))  # dict of float dict
            for block_id_key, items in self.block_item_type_counts.items():
                block_total = 0
                for item_key, item_count in items.items():
                    block_total += item_count
                for item_key, item_count in items.items():
                    self.block_item_type_percentages[block_id_key][item_key] = self.block_item_type_counts[block_id_key][item_key] / block_total

            # Calculate the session-level trial type counts from the block-level counts
            self.session_trial_type_counts = defaultdict(int)
            for block_id_key, items in self.block_trial_type_counts.items():
                for item_key, item_count in items.items():
                    self.session_trial_type_counts[item_key] += item_count

            # Calculate the session-level trial type percentages
            self.session_trial_type_percentages = defaultdict(float)
            for trial_type, trial_type_count in self.session_trial_type_counts.items():
                self.session_trial_type_percentages[trial_type] = trial_type_count / self.trial_count

            # Calculate the session-level item type counts from the block-level counts
            self.session_item_type_counts = defaultdict(int)
            for block_id_key, items in self.block_item_type_counts.items():
                for item_key, item_count in items.items():
                    self.session_item_type_counts[item_key] += item_count

            # Calculate the session-level item type percentages
            self.session_item_type_percentages = defaultdict(float)
            for item_type, item_type_count in self.session_item_type_counts.items():
                self.session_item_type_percentages[item_type] = item_type_count / self.trial_count

        def count_raw_events():
            # "Raw data"
            raw_events = self.get_keypath_value(row, 'data.0.rawEvents')
            for raw_event in raw_events:
                self.raw_count['on'][raw_event['eventOn']] += 1
                self.raw_count['off'][raw_event['eventOff']] += 1

        def check_value_labels():
            session_events = self.get_session_events(row)

            # Record healthy/non-healthy label allocation counts
            self.label_allocation_counts = defaultdict(lambda: defaultdict(int))  # dict of int dict
            self.label_allocation_counts['HEALTHY']['1_'] = 0
            self.label_allocation_counts['HEALTHY']['2_'] = 0
            self.label_allocation_counts['NON_HEALTHY']['1_'] = 0
            self.label_allocation_counts['NON_HEALTHY']['2_'] = 0
            for session_event in session_events:
                item_id = session_event['itemID']
                item_type = session_event['itemType']
                for prefix in ['1_', '2_']:
                    if item_id.startswith(prefix):
                        self.label_allocation_counts[item_type][prefix] += 1

            # Record healthy/non-healthy label allocation percentages
            self.label_allocation_item_id_percentages = defaultdict(lambda: defaultdict(float))  # dict of int dict
            healthy_sum = self.label_allocation_counts['HEALTHY']['1_'] + self.label_allocation_counts['HEALTHY']['2_']
            pprint(self.label_allocation_counts)
            non_healthy_sum = self.label_allocation_counts['NON_HEALTHY']['1_'] + self.label_allocation_counts['NON_HEALTHY']['2_']
            self.label_allocation_item_id_percentages['HEALTHY']['1_'] = self.label_allocation_counts['HEALTHY']['1_'] / self.denominator(healthy_sum)
            self.label_allocation_item_id_percentages['HEALTHY']['2_'] = self.label_allocation_counts['HEALTHY']['2_'] / self.denominator(healthy_sum)
            self.label_allocation_item_id_percentages['NON_HEALTHY']['1_'] = self.label_allocation_counts['NON_HEALTHY']['1_'] / self.denominator(non_healthy_sum)
            self.label_allocation_item_id_percentages['NON_HEALTHY']['2_'] = self.label_allocation_counts['NON_HEALTHY']['2_'] / self.denominator(non_healthy_sum)
            total_sum = healthy_sum + non_healthy_sum
            self.label_allocation_item_type_percentages = defaultdict(float)
            self.label_allocation_item_type_percentages['HEALTHY'] = healthy_sum / self.denominator(total_sum)
            self.label_allocation_item_type_percentages['NON_HEALTHY'] = non_healthy_sum / self.denominator(total_sum)

            # Record the item IDs for each value of selected (MB/random/user/upload/non-food)
            self.selected_item_ids = defaultdict(set)  # dict of set
            for session_event in session_events:
                selected = session_event['selected']
                item_id = session_event['itemID']
                self.selected_item_ids[selected].add(item_id)

            # Record the block-level set of unique item IDs
            self.block_item_ids = defaultdict(set)  # dict of set
            for session_event in session_events:
                block_id = session_event['roundID']
                item_id = session_event['itemID']
                self.block_item_ids[block_id].add(item_id)

            # Record the session-level set of unique item IDs
            self.session_item_ids = set()
            for _, item_ids in self.block_item_ids.items():
                self.session_item_ids.update(item_ids)

        def calculate_dependent_variables():
            self.dv_correct_counts = defaultdict(lambda: defaultdict(int))
            session_events = self.get_keypath_value(row, 'data.0.sessionEvents')
            trial_count = len(session_events)
            for session_event in session_events:
                # GO/STOP
                if 'tapResponseType' in session_event:
                    tap_response_type = session_event['tapResponseType']
                    if tap_response_type == 'CORRECT_GO' or tap_response_type == 'CORRECT_STOP':
                        block_id = session_event['roundID']
                        self.dv_correct_counts[block_id][tap_response_type] += 1

            # Calculate the CORRECT_GO/STOP block-level percentages
            self.dv_correct_block_percentages = defaultdict(lambda: defaultdict(int))
            for block_id_key, tap_response_types in self.dv_correct_counts.items():
                block_total = 0
                for tap_response_type, count in tap_response_types.items():
                    block_total += count
                for tap_response_type, count in tap_response_types.items():
                    self.dv_correct_block_percentages[block_id_key][tap_response_type] = count / block_total

            # Calculate the CORRECT_GO/STOP session-level percentages
            dv_correct_session_counts = defaultdict(int)
            self.dv_correct_session_percentages = defaultdict(float)
            for block_id_key, tap_response_types in self.dv_correct_counts.items():
                for tap_response_type_key, item_count in tap_response_types.items():
                    dv_correct_session_counts[tap_response_type_key] += item_count
            correct_total = 0
            for tap_response_type_key, count in dv_correct_session_counts.items():
                correct_total += count
            for tap_response_type_key, count in dv_correct_session_counts.items():
                self.dv_correct_session_percentages[tap_response_type_key] = count / correct_total

            self.dv_correct_go_responses = list()
            self.dv_correct_stop_responses = list()
            self.dv_correct_responses = defaultdict(lambda: defaultdict(list))
            self.dv_incorrect_healthy_selected_responses = list()
            self.dv_incorrect_healthy_not_selected_responses = list()
            self.dv_incorrect_unhealthy_selected_responses = list()
            self.dv_incorrect_unhealthy_not_selected_responses = list()
            for session_event in session_events:
                # GO/STOP
                if 'tapResponseType' in session_event:
                    tap_response_type = session_event['tapResponseType']
                    tap_response_start = self.numericify(session_event['tapResponseStart'])
                    item_type = session_event['itemType']
                    self.dv_correct_responses[tap_response_type][item_type].append(tap_response_start)
                    if tap_response_type == 'CORRECT_GO':
                        self.dv_correct_go_responses.append(tap_response_start)
                    if tap_response_type == 'CORRECT_STOP':
                        self.dv_correct_stop_responses.append(tap_response_start)

                    trial_type = session_event['trialType']
                    selected = session_event['selected']
                    if trial_type == 'STOP' and tap_response_type != 'CORRECT_STOP':
                        if item_type == 'HEALTHY':
                            if selected != 'random':
                                self.dv_incorrect_healthy_selected_responses.append(tap_response_start)
                            if selected == 'random':
                                self.dv_incorrect_healthy_not_selected_responses.append(tap_response_start)
                        if item_type == 'NON_HEALTHY':
                            if selected != 'random':
                                self.dv_incorrect_unhealthy_selected_responses.append(tap_response_start)
                            if selected == 'random':
                                self.dv_incorrect_unhealthy_not_selected_responses.append(tap_response_start)

        calculate_durations()
        count_trial_and_types()
        count_raw_events()
        check_value_labels()
        self.check_points(row)
        calculate_dependent_variables()
        self.calculate_ssrt(row)

        spreadsheet = Spreadsheet()

        def cell(row, column):
            return '{}{}'.format(str(row), str(column))

        A = 'A'
        B = 'B'
        C = 'C'
        D = 'D'
        E = 'E'
        F = 'F'
        G = 'G'
        H = 'H'
        I = 'I'
        J = 'J'
        letters = [A, B, C, D, E, F, G, H, I, J]
        wb = Workbook()

        # Trial Counts
        spreadsheet.select_sheet('Trial Count')
        spreadsheet.set_values(['Trial Count', self.trial_count], advance_by_rows=2)

        # Raw Counts
        spreadsheet.set_value('Raw Counts', advance_by_rows=2)
        spreadsheet.set_values(['On'])
        for key, value in self.raw_count['on'].items():
            spreadsheet.set_values([key, value])
        spreadsheet.advance_row()
        spreadsheet.set_values(['Off'])
        for key, value in self.raw_count['off'].items():
            spreadsheet.set_values([key, value])

        # Durations
        print('\nTRIAL DURATIONS:')
        for key in self.durations:
            print(key, self.durations[key])

        # Trial Stats
        spreadsheet.select_sheet('Stats')
        spreadsheet.set_values(['Field', 'Min', 'Max', 'Mean', 'St Dev'])
        for field, stats in self.trial_stats.items():
            spreadsheet.set_values([field, stats['min'], stats['max'], stats['mean'], stats['stdev']])

        # Trial Types
        spreadsheet.select_sheet('Trial Types')
        # - Blocks
        spreadsheet.set_values(['Session'])
        spreadsheet.set_values(['Trial Type', 'Count', 'Percentage'])
        for trial_type in ['GO', 'STOP']:
            spreadsheet.set_values([
                trial_type,
                self.session_trial_type_counts[trial_type],
                self.session_trial_type_percentages[trial_type]
            ])
        # - Session
        spreadsheet.advance_row()
        spreadsheet.set_values(['Block'])
        spreadsheet.set_values(['Block', 'Trial Type', 'Count', 'Percentage'])
        for block_key in irange(1, 4):
            for trial_type in ['GO', 'STOP']:
                spreadsheet.set_values([
                    block_key,
                    trial_type,
                    self.block_trial_type_counts[block_key][trial_type],
                    self.block_trial_type_percentages[block_key][trial_type]
                ])

        print('\nSESSION TRIAL TYPE COUNTS:')
        pprint(self.session_trial_type_counts)
        print('\nSESSION TRIAL TYPE PERCENTAGES:')
        pprint(self.session_trial_type_percentages)
        print('\nBLOCK TRIAL TYPE COUNTS:')
        pprint(self.block_trial_type_counts)
        print('\nBLOCK TRIAL TYPE PERCENTAGES:')
        pprint(self.block_trial_type_percentages)

        # Trial Items
        # - Block
        r = 1
        trial_types_sheet = wb.create_sheet('Item Types')
        trial_types_sheet[cell(A, r)] = 'Session'
        r += 1
        trial_types_sheet[cell(A, r)] = 'Item Type'
        trial_types_sheet[cell(B, r)] = 'Count'
        trial_types_sheet[cell(C, r)] = 'Percentage'
        for trial_type in ['HEALTHY', 'HEALTHY_RANDOM', 'HEALTHY_NOT_RANDOM', 'NON_HEALTHY']:
            r += 1
            trial_types_sheet[cell(A, r)] = trial_type
            trial_types_sheet[cell(B, r)] = self.session_item_type_counts[trial_type]
            trial_types_sheet[cell(C, r)] = self.session_item_type_percentages[trial_type]
        # - Session
        r += 2
        trial_types_sheet[cell(A, r)] = 'Block'
        r += 1
        trial_types_sheet[cell(A, r)] = 'Block'
        trial_types_sheet[cell(B, r)] = 'Item Type'
        trial_types_sheet[cell(C, r)] = 'Count'
        trial_types_sheet[cell(D, r)] = 'Percentage'
        for block_key in irange(1, 4):
            for trial_type in ['HEALTHY', 'HEALTHY_RANDOM', 'HEALTHY_NOT_RANDOM', 'NON_HEALTHY']:
                r += 1
                trial_types_sheet[cell(A, r)] = block_key
                trial_types_sheet[cell(B, r)] = trial_type
                trial_types_sheet[cell(C, r)] = self.block_item_type_counts[block_key][trial_type]
                trial_types_sheet[cell(D, r)] = self.block_item_type_percentages[block_key][trial_type]
        print('\nSESSION ITEM TYPE COUNTS:')
        pprint(self.session_item_type_counts)
        print('\nSESSION ITEM TYPE PERCENTAGES:')
        pprint(self.session_item_type_percentages)
        print('\nBLOCK ITEM TYPE COUNTS:')
        pprint(self.block_item_type_counts)
        print('\nBLOCK ITEM TYPE PERCENTAGES:')
        pprint(self.block_item_type_percentages)

        # Raw Rounds
        spreadsheet.select_sheet('Raw Rounds')
        spreadsheet.set_values(['Round', 'Count'])
        for block_key in irange(1, 4):
            spreadsheet.set_values([
                block_key,
                len(self.raw_round_trial_counts[block_key])
            ])

        # Label Allocation Counts
        r = 1
        label_allocation_count_sheet = wb.create_sheet('Label Allocations')
        label_allocation_count_sheet[cell(A, r)] = 'Session'
        r += 1
        label_allocation_count_sheet[cell(A, r)] = 'Trial Type'
        label_allocation_count_sheet[cell(B, r)] = 'Prefix'
        label_allocation_count_sheet[cell(C, r)] = 'Count'
        label_allocation_count_sheet[cell(D, r)] = 'Percentage'
        for item_type in ['HEALTHY', 'NON_HEALTHY']:
            for prefix in ['1_', '2_']:
                r += 1
                label_allocation_count_sheet[cell(A, r)] = item_type
                label_allocation_count_sheet[cell(B, r)] = prefix
                label_allocation_count_sheet[cell(C, r)] = self.label_allocation_counts[item_type][prefix]
                label_allocation_count_sheet[cell(D, r)] = self.label_allocation_item_id_percentages[item_type][prefix]
        print('\nLABEL ALLOCATION COUNTS:')
        pprint(self.label_allocation_counts)

        # Label Allocation Percentages
        print('\nLABEL ALLOCATION PERCENTAGES:')
        pprint(self.label_allocation_item_id_percentages)
        pprint(self.label_allocation_item_type_percentages)

        r += 2
        label_allocation_count_sheet[cell(A, r)] = 'Item Type'
        label_allocation_count_sheet[cell(B, r)] = 'Percentage'
        r += 1
        for item_type in self.label_allocation_item_type_percentages:
            label_allocation_count_sheet[cell(A, r)] = item_type
            label_allocation_count_sheet[cell(B, r)] = self.label_allocation_item_type_percentages[item_type]
            r += 1

        # Selected Item IDs
        r = 1
        selected_items_sheet = wb.create_sheet('Selected Items')
        for index, selected_label in enumerate(self.selected_item_ids):
            selected_items_sheet[cell(letters[index], r)] = selected_label
        r += 1
        for index, selected_label in enumerate(self.selected_item_ids):
            for i, selected_item_id in enumerate(self.selected_item_ids[selected_label]):
                selected_items_sheet[cell(letters[index], r+i)] = selected_item_id
        print('\nSELCTED ITEM IDs:')
        print(self.selected_item_ids)

        # Unique Item IDs
        # - Session
        spreadsheet.select_sheet('Unique Items - Session')
        for index, unique_item_id in enumerate(self.session_item_ids):
            spreadsheet.set_values([unique_item_id])
        # - Blocks
        spreadsheet.select_sheet('Unique Items - Blocks')
        for block_key in irange(1, 4):
            spreadsheet.set_value(block_key)
        spreadsheet.advance_row()
        for block_key in irange(1, 4):
            for index, unique_item_id in enumerate(self.block_item_ids[block_key]):
                spreadsheet.column = block_key
                spreadsheet.set_value(unique_item_id, cell_offset=(block_key, index+1))

        print('\nSESSION ITEM IDs:')
        print(self.session_item_ids)
        print('\nBLOCK ITEM IDs:')
        pprint(self.block_item_ids)

        r = 1
        correct_count_sheet = wb.create_sheet('Correct Counts')
        correct_count_sheet[cell(A, r)] = 'Block'
        r += 1
        correct_count_sheet[cell(A, r)] = 'Block'
        correct_count_sheet[cell(B, r)] = 'Trial Type'
        correct_count_sheet[cell(C, r)] = 'Count'
        correct_count_sheet[cell(C, r)] = 'Count'
        r += 1
        for block_key in irange(1, 4):
            for trial_type in ['CORRECT_GO', 'CORRECT_STOP']:
                correct_count_sheet[cell(A, r)] = block_key
                correct_count_sheet[cell(B, r)] = trial_type
                correct_count_sheet[cell(C, r)] = self.dv_correct_counts[block_key][trial_type]
                r += 1


        print('\nDV CORRECT COUNTS:')
        pprint(self.dv_correct_counts)
        print('DV BLOCK LEVEL CORRECT PERCENTAGES:')
        pprint(self.dv_correct_block_percentages)
        print('DV SESSION LEVEL CORRECT PERCENTAGES:')
        pprint(self.dv_correct_session_percentages)
        # assert False

        print('mean CORRECT_GO responses', self.mean(self.dv_correct_go_responses))
        print('mean CORRECT_GO HEALTHY responses', self.mean(self.dv_correct_responses['CORRECT_GO']['HEALTHY']))
        print('mean CORRECT_GO UNHEALTHY responses', self.mean(self.dv_correct_responses['CORRECT_GO']['NON_HEALTHY']))
        print('mean CORRECT_STOP responses', self.mean(self.dv_correct_stop_responses))
        print('mean CORRECT_STOP HEALTHY responses', self.mean(self.dv_correct_responses['CORRECT_STOP']['HEALTHY']))
        print('mean CORRECT_STOP UNHEALTHY responses', self.mean(self.dv_correct_responses['CORRECT_STOP']['NON_HEALTHY']))

        print('mean INCORRECT HEALTHY SELECTED responses', self.mean(self.dv_incorrect_healthy_selected_responses))
        print('mean INCORRECT HEALTHY NOT SELECTED responses', self.mean(self.dv_incorrect_healthy_not_selected_responses))
        print('mean INCORRECT UNHEALTHY SELECTED responses', self.mean(self.dv_incorrect_unhealthy_selected_responses))
        print('mean INCORRECT UNHEALTHY NOT SELECTED responses', self.mean(self.dv_incorrect_unhealthy_not_selected_responses))

        excel_filename = 'DerivedValues.xlsx'
        wb.save(excel_filename)
        spreadsheet.save('new-DerivedValues.xlsx')
        assert False
        # print('\nLog')

        self.session_event_log.print()


class StopDataExtractor(AbstractStopDataExtractor):

    type = 'STOP'


class RestraintDataExtractor(AbstractStopDataExtractor):

    type = 'RESTRAINT'


class NAStopDataExtractor(AbstractStopDataExtractor):

    type = 'NASTOP'


class NARestraintDataExtractor(AbstractStopDataExtractor):

    type = 'NARESTRAINT'


class GStopDataExtractor(AbstractStopDataExtractor):

    type = 'GSTOP'


class GRestraintDataExtractor(AbstractStopDataExtractor):

    type = 'GRESTRAINT'


class DoubleDataExtractor(AbstractStopDataExtractor):

    type = 'DOUBLE'

    def check_points(self, row):

        def check_go(initial_tap_response_type, second_tap_response_type):
            if initial_tap_response_type == 'CORRECT_GO' and second_tap_response_type == 'N/A':
                check_passed = points_this_trial == 20
                self.session_event_log.log_if_check_failed(check_passed, session_event)
            elif initial_tap_response_type == 'INCORRECT_GO' and second_tap_response_type == 'N/A':
                check_passed = points_this_trial == -20
                self.session_event_log.log_if_check_failed(check_passed, session_event)
            else:
                check_passed = points_this_trial == -50
                self.session_event_log.log_if_check_failed(check_passed, session_event)

        def check_double(initial_tap_response_type, second_tap_response_type):
            if initial_tap_response_type == 'CORRECT' and second_tap_response_type == 'CORRECT':
                check_passed = points_this_trial == 50
                self.session_event_log.log_if_check_failed(check_passed, session_event)
            else:
                check_passed = points_this_trial == -50
                self.session_event_log.log_if_check_failed(check_passed, session_event)

        checks = {
            'GO': check_go,
            'DOUBLE': check_double,
        }

        running_total = 0
        session_events = self.get_session_events(row)
        for session_event in session_events:
            # Points Awarded
            trial_type = session_event['trialType']
            initial_tap_response_type = session_event['initialTapResponseType']
            second_tap_response_type = session_event['secondTapResponseType']
            points_this_trial = session_event['pointsThisTrial']
            checks[trial_type](initial_tap_response_type, second_tap_response_type)
            # Running Total
            running_total += points_this_trial
            points_running_total = session_event['pointsRunningTotal']
            check_passed = points_running_total == running_total
            self.session_event_log.log_if_check_failed(check_passed, session_event, extra_message='points_running_total != running_total')

    def check_tap_responses(self, row):
        # trs = tap response start
        initial_tap_response_checks = {
            'GO': {
                'CORRECT_GO': lambda trs, session_event: trs > 0 and self.within_first_stimulus_boundary(session_event),
                'INCORRECT_GO': lambda trs, session_event: trs == 0,
                'MISS_GO': lambda trs, session_event: trs > 0 and self.outside_first_stimulus_boundary(session_event),
            },
            'DOUBLE': {
                'CORRECT': lambda trs, session_event: trs == 0,
                'INCORRECT': lambda trs, session_event: trs > 0 and self.within_first_stimulus_boundary(session_event),
                'MISS': lambda trs, session_event: trs > 0 and self.outside_first_stimulus_boundary(session_event),
            }
        }
        second_tap_response_checks = {
            'GO': {
                'N/A': lambda trs, session_event: trs == 0,
                'INCORRECT_DOUBLE_GO': lambda trs, session_event: trs > 0 and self.within_second_stimulus_boundary(session_event),
                'MISS_GO': lambda trs, session_event: trs > 0 and self.outside_second_stimulus_boundary(session_event),
            },
            'DOUBLE': {
                'CORRECT': lambda trs, session_event: trs > 0 and self.within_second_stimulus_boundary(session_event),
                'INCORRECT': lambda trs, session_event: trs == 0,
                'MISS': lambda trs, session_event: trs > 0 and self.outside_second_stimulus_boundary(session_event),
            }
        }

        def check_tap_response(tap_response_checks, session_event, prefix):
            tap_response_type = session_event['{}TapResponseType'.format(prefix)]
            # Fix incorrectly named key
            if tap_response_type == 'INCORR_DOUB_GO':
                tap_response_type = 'INCORRECT_DOUBLE_GO'
            tap_response_start = self.numericify(session_event['{}TapResponseStart'.format(prefix)])
            check_result = tap_response_checks[trial_type][tap_response_type](tap_response_start, session_event)
            self.session_event_log.log_if_check_failed(check_result, session_event)

        session_events = self.get_session_events(row)
        for session_event in session_events:
            trial_type = session_event['trialType']
            check_tap_response(initial_tap_response_checks, session_event, 'initial')
            check_tap_response(second_tap_response_checks, session_event, 'second')

    def within_first_stimulus_boundary(self, session_event):
        self.within_stimulus_boundary(session_event, prefix='initialTapResponsePosition')

    def outside_first_stimulus_boundary(self, session_event):
        self.outside_stimulus_boundary(session_event, prefix='initialTapResponsePosition')

    def within_second_stimulus_boundary(self, session_event):
        self.within_stimulus_boundary(session_event, prefix='secondTapResponsePosition')

    def outside_second_stimulus_boundary(self, session_event):
        self.outside_stimulus_boundary(session_event, prefix='secondTapResponsePosition')

    def calculate_ssrt(self, row):
        pass

    def calculate(self, row):
        super(DoubleDataExtractor, self).calculate(row)
