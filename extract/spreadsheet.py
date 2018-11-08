import string
from openpyxl import Workbook


class Spreadsheet:
    """A wrapper around the openpyxl Workbook that provides a typewriter style write-and-advance metaphor"""

    alphabet = list(string.ascii_uppercase)

    def __init__(self):
        self.row = 1
        self.column = 1
        self.workbook = Workbook()

    def letter(self, index):
        return self.alphabet[index]

    def current_cell(self):
        """
        Return the Excel letter/number coordinates of the current cell using the
        1-based row,column coordinates of the current cell, e.g. 3,2 -> C2
        """
        return '{}{}'.format(self.letter(self.column-1), str(self.row))

    def select_sheet(self, sheet_name):
        """Set the active sheet by the given name (creating it if necessary)"""
        sheet = self.get_sheet(sheet_name)
        self.workbook.active = sheet

    def get_sheet(self, sheet_name):
        """Return an existing sheet with the given name or create a new sheet with that name"""
        if sheet_name in self.workbook.sheetnames:
            return self.workbook.get_sheet_by_name(sheet_name)
        else:
            return self.workbook.create_sheet(sheet_name)

    def current_sheet(self):
        return self.workbook.active

    def set_value(self, value, advance_row=False, advance_by_rows=1):
        """Set the value of the current cell"""
        sheet = self.current_sheet()
        cell = self.current_cell()
        sheet[cell] = value
        if advance_row or advance_by_rows > 1:
            for _ in range(0, advance_by_rows):
                self.advance_row()
        else:
            self.advance_column()

    def advance_column(self):
        self.column += 1

    def advance_row(self):
        self.column = 1
        self.row += 1

    def save(self, filename):
        self.workbook.save(filename)


if __name__ == '__main__':
    spreadsheet = Spreadsheet()
    print(spreadsheet.current_cell())
    spreadsheet.advance_column()
    print(spreadsheet.current_cell())
    spreadsheet.advance_row()
    print(spreadsheet.current_cell())
    spreadsheet.select_sheet('Stats')
    print(spreadsheet.current_sheet().title)
