import json

import os
import pprint
import statistics

import sys
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from openpyxl.cell import Cell


class ReadStop:

    def __init__(self):
        self.filename = ''
        self.all_data_file_idx = 0
        self.summary_count = 1
        self.session_durations = []
        self.trial_type_counts = {}

    def read_stop_file(self, json_file):
        with open(json_file, 'r') as json_file_handle:
            json_blob = json.load(json_file_handle)
            self.output_srgame_spreadsheet(json_blob[0])

    def resize_sheet_columns(self, sheet, row_number=0):
        # Changing the column width to approx the length of the headers
        for idx, col_header in enumerate(list(sheet.rows)[row_number]):
            dim = sheet.column_dimensions[get_column_letter(idx + 1)]
            dim.width = len(col_header.value)

    def output_srgame_spreadsheet(self, json_blob, folder='test_folder'):
        # extra_headers = [{
        #     'col': 'tapResponseStart',
        #     'output': 'avg_response',
        #     'method': 'avg'
        # }]

        output_filepath = '{}_{}_{}.xlsx'.format(
            json_blob['userId'],
            json_blob['type'],
            json_blob['data'][0]['gameSessionID']
        )

        session_data = json_blob['data'][0]['sessionEvents']
        col_headers = sorted(session_data[0].keys())

        wb = Workbook()
        self.wb = wb
        # wb.sheetnames
        sheet = wb.active
        sheet.title = 'Data Output'
        print('\ncreating data output sheet for {}\n'.format(output_filepath))

        self.summary_count = 0
        response_times = []
        session_start = json_blob['data'][0]['sessionStart']
        session_end = json_blob['data'][0]['sessionEnd']
        previous_trial_start = None

        trial_durations = []
        stop_signal_durations = []
        stimulus_durations = []
        signal_stop_delays = []
        interTrialIntervals = []
        stim_stop_sig_offset_diffs = []

        healthy_ones = 0
        healthy_twos = 0
        non_healthy_ones = 0
        non_healthy_twos = 0

        itemID_selected = {
            'MB': [],
            'random': [],
            'user': [],
            'upload': [],
            'non-food': []
        }

        max_round = 0
        round_trial_counts = {}

        self.trial_type_counts = {
            'total': {
                'STOP': 0,
                'GO': 0,
                'HEALTHY': 0,
                'NON_HEALTHY': 0,
                'HEALTHY_RANDOM': 0,
                'HEALTHY_NOT_RANDOM': 0
            },
            'blocks': {}
        }

        sheet.cell(
            column=1, row=1, value='Data ({}) from {}'.format(
                self.all_data_file_idx,
                self.filename)
        )

        for idx, col_header in enumerate(col_headers):
            sheet.cell(column=idx+1, row=2, value=col_header)

        # Rows within a session
        for s_idx, trial in enumerate(session_data):

            # print('row', s_idx)

            for col_idx, col_header in enumerate(col_headers):
                value = trial[col_header]

                if col_header == 'tapResponseStart' and value:
                    response_times.append(value)

                sheet.cell(column=col_idx+1, row=s_idx+3, value=value)

            # roundID
            if 'roundID' in trial and trial['roundID']:
                if trial['roundID'] > max_round:
                    max_round = trial['roundID']

                if trial['roundID'] in round_trial_counts:
                    round_trial_counts[trial['roundID']] = round_trial_counts[trial['roundID']] + 1
                else:
                    round_trial_counts[trial['roundID']] = 0

            # Trial duration: trialEnd - trialStart
            sheet.cell(column=len(col_headers) + 1, row=2, value="Trial Duration")
            if 'trialEnd' in trial and 'trialStart' in trial:
                if trial['trialEnd'] and trial['trialStart']:
                    trial_duration = trial['trialEnd'] - trial['trialStart']
                    sheet.cell(column=len(col_headers) + 1, row=s_idx+3, value=trial_duration)
                    trial_durations.append(trial_duration)

            # Stop signal duration: stopSignalOffset - stopSignalOnset
            sheet.cell(column=len(col_headers) + 2, row=2, value="Stop Signal Duration")
            if 'stopSignalOffset' in trial and 'stopSignalOnset' in trial:
                if trial['stopSignalOffset'] and trial['stopSignalOnset']:
                    stop_signal_duration = trial['stopSignalOffset'] - trial['stopSignalOnset']
                    sheet.cell(column=len(col_headers) + 2, row=s_idx+3, value=stop_signal_duration)
                    stop_signal_durations.append(stop_signal_duration)

            # Stimulus duration: stimulusOffset - stimulusOnset
            sheet.cell(column=len(col_headers) + 3, row=2, value="Stimulus Duration")
            if 'stimulusOffset' in trial and 'stimulusOnset' in trial:
                stimulus_duration = trial['stimulusOffset'] - trial['stimulusOnset']
                sheet.cell(column=len(col_headers) + 3, row=s_idx+3, value=stimulus_duration)
                stimulus_durations.append(stimulus_duration)

            # Difference between signal onset and stop signal delay: stopSignalOnset - stopSignalDelay
            sheet.cell(column=len(col_headers) + 4, row=2, value="stopSignalOnset - stopSignalDelay")
            if 'stopSignalOnset' in trial and 'stopSignalDelay' in trial:
                if trial['stopSignalOnset'] and trial['stopSignalDelay']:
                    signal_stop_delay = trial['stopSignalOnset'] - trial['stopSignalDelay']
                    sheet.cell(column=len(col_headers) + 4, row=s_idx + 3, value=signal_stop_delay)
                    signal_stop_delays.append(signal_stop_delay)

            # Duration between signal offset and stimulus offset: stimulusOffset - stopSignalOffset
            sheet.cell(column=len(col_headers) + 5, row=2, value="stimulusOffset - stopSignalOffset")
            if 'stimulusOffset' in trial and 'stopSignalOffset' in trial:
                if trial['stimulusOffset'] and trial['stopSignalOffset']:
                    # TODO this is zero apparently...
                    stim_stop_sig_offset_diff = trial['stimulusOffset'] - trial['stopSignalOffset']
                    sheet.cell(column=len(col_headers) + 5, row=s_idx + 3, value=stim_stop_sig_offset_diff)
                    stim_stop_sig_offset_diffs.append(stim_stop_sig_offset_diff)

            # interTrialInterval
            # Duration between trials: trialOnsetB - trialOnsetA,
            # where B is the trial of interest and A is the trial preceding that.
            # Calculations needed without 1st trial of each block (because of inter-block/graphical feedback).
            # TODO does 'trialOnsetA' mean trialStart?
            if 'trialStart' in trial:
                if trial['trialStart']:
                    if previous_trial_start:
                        interTrialInterval = trial['trialStart'] - previous_trial_start
                        sheet.cell(column=len(col_headers) + 6, row=2,
                                   value="interTrialInterval")
                        sheet.cell(column=len(col_headers) + 6, row=s_idx + 3,
                                   value=interTrialInterval)
                        interTrialIntervals.append(interTrialInterval)
                    previous_trial_start = trial['trialStart']

            if 'trialType' in trial:
                # Global store of numbers of trialType (STOP, GO) by Block
                self.store_trial_round_game_type(trial)

                # Keep a count of trialTypes, probably STOP or GO
                if trial['trialType'] not in self.trial_type_counts['total']:
                    self.trial_type_counts['total'][trial['trialType']] = 1
                else:
                    self.trial_type_counts['total'][trial['trialType']] += 1

            if 'itemType' in trial:
                if trial['itemType'] == 'HEALTHY':
                    self.trial_type_counts['total']['HEALTHY'] += 1

                    if trial['selected'] == 'RANDOM':
                        self.trial_type_counts['total']['HEALTHY_RANDOM'] += 1
                    else:
                        self.trial_type_counts['total']['HEALTHY_NOT_RANDOM'] += 1

                    # the number (and %) of foods labelled ‘HEALTHY’ (itemType)
                    # for which itemID starts with ‘1_ and ‘2_
                    if 'itemID' in trial:
                        if str(trial['itemID']).startswith('1_'):
                            healthy_ones += 1
                        if str(trial['itemID']).startswith('2_'):
                            healthy_twos += 1

                if trial['itemType'] == 'NON_HEALTHY':
                    self.trial_type_counts['total']['NON_HEALTHY'] += 1

                    # the number (and %) of foods labelled ‘HEALTHY’ (itemType)
                    # for which itemID starts with ‘1_ and ‘2_
                    if 'itemID' in trial:
                        if str(trial['itemID']).startswith('1_'):
                            non_healthy_ones += 1
                        if str(trial['itemID']).startswith('2_'):
                            non_healthy_twos += 1

            if 'selected' in trial:
                itemID_selected[trial['selected']].append(trial['itemID'])

        # # Changing the column width to approx the length of the headers
        # for idx, col_header in enumerate(list(sheet.rows)[1]):
        #     dim = sheet.column_dimensions[get_column_letter(idx + 1)]
        #     dim.width = len(col_header.value)
        self.resize_sheet_columns(sheet, 1)

        try:
            print('\ncreating summary sheet for {}\n'.format(output_filepath))

            wb.create_sheet('Data Summary')
            sheet2 = wb['Data Summary']

            sheet2.cell(column=1, row=2, value='Response Avg')
            sheet2.cell(column=2, row=2, value='Response Low')
            sheet2.cell(column=3, row=2, value='Response High')

            # If we don't have any response times, put N/A to maintain data shape
            if len(response_times):
                response_times = sorted(response_times)
                avg_res = sum(response_times) / len(response_times)
                sheet2.cell(column=1, row=3, value=avg_res)
                sheet2.cell(column=2, row=3, value=response_times[0])
                sheet2.cell(column=3, row=3, value=response_times[-1])

                self.store_summary_key_value('Response Avg', avg_res)
                self.store_summary_key_value('Response Low', response_times[0])
                self.store_summary_key_value('Response High', response_times[-1])

            else:
                sheet2.cell(column=1, row=3, value='N/A')
                sheet2.cell(column=2, row=3, value='N/A')
                sheet2.cell(column=3, row=3, value='N/A')

                self.store_summary_key_value('Response Avg', 'N/A')
                self.store_summary_key_value('Response Low', 'N/A')
                self.store_summary_key_value('Response High', 'N/A')

            sheet2.cell(column=4, row=2, value='Session Duration')
            session_duration = (session_end - session_start)
            self.session_durations.append(session_duration)
            print('self.session_durations', self.session_durations)
            sheet2.cell(column=4, row=3, value=session_duration)
            self.store_summary_key_value('Session Duration', session_duration)

            sheet2.cell(column=5, row=2, value='Session Event Count')
            sheet2.cell(column=5, row=3, value=len(session_data))
            self.store_summary_key_value('Session Event Count', len(session_data))

            sheet2.cell(column=6, row=2, value='Max round ID')
            sheet2.cell(column=6, row=3, value=max_round)
            self.store_summary_key_value('Max round ID', max_round)

            # round number and count
            sheet2.cell(column=7, row=2, value='Round trial numbers')
            sheet2.cell(column=8, row=2, value='Round trial count')
            for itr, value in enumerate(round_trial_counts):
                sheet2.cell(column=7, row=3 + itr, value=value)
                sheet2.cell(column=8, row=3 + itr, value=round_trial_counts[value])

            sheet2.cell(column=9, row=2, value='Num GOs')
            sheet2.cell(column=9, row=3, value=self.trial_type_counts['total']['GO'])
            self.store_summary_key_value('Num GOs', self.trial_type_counts['total']['GO'])

            sheet2.cell(column=10, row=2, value='Num STOPs')
            sheet2.cell(column=10, row=3, value=self.trial_type_counts['total']['STOP'])
            self.store_summary_key_value('Num STOPs', self.trial_type_counts['total']['STOP'])

            sheet2.cell(column=11, row=2, value='Num HEALTHYs')
            sheet2.cell(column=11, row=3, value=self.trial_type_counts['total']['HEALTHY'])
            self.store_summary_key_value('Num HEALTHYs', self.trial_type_counts['total']['HEALTHY'])

            sheet2.cell(column=12, row=2, value='Num NON_HEALTHYs')
            sheet2.cell(column=12, row=3, value=self.trial_type_counts['total']['NON_HEALTHY'])
            self.store_summary_key_value('Num NON_HEALTHYs', self.trial_type_counts['total']['NON_HEALTHY'])

            # HEALTHY_RANDOM
            sheet2.cell(column=13, row=2, value='Num HEALTHY_RANDOMs')
            sheet2.cell(column=13, row=3, value=self.trial_type_counts['total']['HEALTHY_RANDOM'])
            self.store_summary_key_value('Num HEALTHY_RANDOMs', self.trial_type_counts['total']['HEALTHY_RANDOM'])

            # HEALTHY_NOT_RANDOM
            sheet2.cell(column=14, row=2, value='Num HEALTHY_NOT_RANDOMs')
            sheet2.cell(column=14, row=3, value=self.trial_type_counts['total']['HEALTHY_NOT_RANDOM'])
            self.store_summary_key_value('Num HEALTHY_NOT_RANDOMs',
                                         self.trial_type_counts['total']['HEALTHY_NOT_RANDOM'])

            for block in self.trial_type_counts['blocks'].keys():
                for key in self.trial_type_counts['blocks'][block].keys():
                    self.store_summary_key_value("{}-{}".format(block, key),
                                                 self.trial_type_counts['blocks'][block][key])
                    if key in self.trial_type_counts['total']:
                        self.store_summary_key_value("{}_block_{} / {}_total".format(key, block, key),
                                                     (self.trial_type_counts['blocks'][block][key]/self.trial_type_counts['total'][key]))

            self.store_summary_key_value('Num HEALTHY 1_s', healthy_ones)
            self.store_summary_key_value('Num HEALTHY 2_s', healthy_twos)
            self.store_summary_key_value('Num Non HEALTHY 1_s', non_healthy_ones)
            self.store_summary_key_value('Num Non HEALTHY 2_s', non_healthy_twos)

            if len(trial_durations):
                min_trial_duration = min(trial_durations)
                max_trial_duration = max(trial_durations)
                mean_trial_duration = statistics.mean(trial_durations)
                trial_duration_stdev = statistics.stdev(trial_durations)

                self.store_summary_key_value('Min Trial Duration', min_trial_duration)
                self.store_summary_key_value('Max Trial Duration', max_trial_duration)
                self.store_summary_key_value('Mean Trial Duration', mean_trial_duration)
                self.store_summary_key_value('Trial Duration Std Dev', trial_duration_stdev)

            if len(stop_signal_durations):
                min_stop_signal_duration = min(stop_signal_durations)
                max_stop_signal_duration = max(stop_signal_durations)
                mean_stop_signal_duration = statistics.mean(stop_signal_durations)
                stop_signal_duration_stdev = statistics.stdev(stop_signal_durations)

                self.store_summary_key_value('Min Stop Signal Duration', min_stop_signal_duration)
                self.store_summary_key_value('Max Stop Signal Duration', max_stop_signal_duration)
                self.store_summary_key_value('Mean Stop Signal Duration', mean_stop_signal_duration)
                self.store_summary_key_value('Stop Signal Duration Std Dev', stop_signal_duration_stdev)

            if len(stimulus_durations):
                min_stimulus_duration = min(stimulus_durations)
                max_stimulus_duration = max(stimulus_durations)
                mean_stimulus_duration = statistics.mean(stimulus_durations)
                stimulus_duration_stdev = statistics.stdev(stimulus_durations)

                self.store_summary_key_value('Min Stimulus Duration', min_stimulus_duration)
                self.store_summary_key_value('Max Stimulus Duration', max_stimulus_duration)
                self.store_summary_key_value('Mean Stimulus Duration', mean_stimulus_duration)
                self.store_summary_key_value('Stop Stimulus Std Dev', stimulus_duration_stdev)

            if len(signal_stop_delays):
                min_signal_stop_delay = min(signal_stop_delays)
                max_signal_stop_delay = max(signal_stop_delays)
                mean_signal_stop_delay = statistics.mean(signal_stop_delays)
                signal_stop_delay_stdev = statistics.stdev(signal_stop_delays)

                self.store_summary_key_value('Min Signal Stop Delay', min_signal_stop_delay)
                self.store_summary_key_value('Max Signal Stop Delay', max_signal_stop_delay)
                self.store_summary_key_value('Mean Signal Stop Delay', mean_signal_stop_delay)
                self.store_summary_key_value('Signal Stop Delay Std Dev', signal_stop_delay_stdev)

            if len(interTrialIntervals):
                min_interTrialInterval = min(interTrialIntervals)
                max_interTrialInterval = max(interTrialIntervals)
                mean_interTrialInterval = statistics.mean(interTrialIntervals)
                interTrialInterval_stdev = statistics.stdev(interTrialIntervals)

                self.store_summary_key_value('Min InterTrial Interval', min_interTrialInterval)
                self.store_summary_key_value('Max InterTrial Interval', max_interTrialInterval)
                self.store_summary_key_value('Mean InterTrial Interval', mean_interTrialInterval)
                self.store_summary_key_value('InterTrial Interval Std Dev', interTrialInterval_stdev)

            if len(stim_stop_sig_offset_diffs):
                min_stim_stop_sig_offset_diff = min(stim_stop_sig_offset_diffs)
                max_stim_stop_sig_offset_diff = max(stim_stop_sig_offset_diffs)
                mean_stim_stop_sig_offset_diff = statistics.mean(stim_stop_sig_offset_diffs)
                stim_stop_sig_offset_diff_stdev = statistics.stdev(stim_stop_sig_offset_diffs)

                self.store_summary_key_value('Min Stimulus - Stop Signal Offset Difference',
                                             min_stim_stop_sig_offset_diff)
                self.store_summary_key_value('Max Stimulus - Stop Signal Offset Difference',
                                             max_stim_stop_sig_offset_diff)
                self.store_summary_key_value('Mean Stimulus - Stop Signal Offset Difference',
                                             mean_stim_stop_sig_offset_diff)
                self.store_summary_key_value('Stimulus - Stop Signal Offset Difference Std Dev',
                                             stim_stop_sig_offset_diff_stdev)

            itemid_set = set()
            for key in itemID_selected:
                self.store_summary_key_value('itemID_selected_{}'.format(key),
                                             str(itemID_selected[key])
                                             # json.dumps(itemID_selected[key], indent=4)
                                             )
                itemid_set.update(itemID_selected[key])
            self.store_summary_key_value('Unique itemIDs', str(list(itemid_set)))

            self.resize_sheet_columns(sheet2, 1)
        except Exception as e1:
            print('Data Summary error', e1)
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)

        parent = os.path.dirname(os.path.abspath(output_filepath))
        new_folder = os.path.join(parent, folder)
        if not os.path.exists(new_folder):
            os.mkdir(new_folder)

        output_file_and_dir = os.path.join(new_folder, output_filepath)
        print(output_file_and_dir)
        wb.save(output_file_and_dir)

        print(pprint.pformat(self.trial_type_counts, indent=4))

    def sort_all_data(self, all_data_file):
        self.filename = all_data_file

        head, tail = os.path.split(self.filename)
        new_folder_name = tail.split('.')[0]

        with open(all_data_file, 'r') as all_data_handle:
            all_data =  json.load(all_data_handle)

            all_types = []
            task_count = 0
            for idx, a in enumerate(all_data):

                # print((idx + 1), a['type'])

                self.all_data_file_idx = idx

                all_types.append(a['type'])
                if 'data' in a:
                    if isinstance(a['data'], list) and len(a['data']):
                        # print(a['data'])
                        if 'sessionEvents' in a['data'][0]:
                            task_count += len(a['data'][0]['sessionEvents'])

                if a['type'] == 'virtual-supermarket-selected':
                    print('virtual-supermarket-selected idx', idx)
                else:

                    # print('\n', type(a['data']), a['data'], '\n')
                    if 'data' in a and isinstance(a['data'], list) and len(a['data']) and 'gameSessionID' in a['data'][0]:

                        # It's a game - we'll inspect further
                        self.output_srgame_spreadsheet(
                            a,
                            folder=new_folder_name
                        )

            print(set(all_types))
            print('task_count', task_count)

    def store_summary_key_value(self, key, value, sheet_title='Data Summary 2'):
        if sheet_title not in self.wb:
            print('creating a sheet', sheet_title)
            self.wb.create_sheet(sheet_title)

        sheet = self.wb[sheet_title]
        self.summary_count += 1

        sheet.cell(column=1,
                   row=self.summary_count,
                   value=key)
        sheet.cell(column=2,
                   row=self.summary_count,
                   value=value)

        dim = sheet.column_dimensions[get_column_letter(1)]
        if not dim.width or len(str(key)) > dim.width:
            dim.width = len(str(key))

        dim = sheet.column_dimensions[get_column_letter(2)]
        if not dim.width or len(str(value)) > dim.width:
            dim.width = len(str(value))

    def store_trial_round_game_type(self, trial):
        # print('\n\n', trial)

        block_number = trial['roundID']
        game_name = trial['trialType']
        # print(block_number, game_name)

        if block_number not in self.trial_type_counts['blocks']:
            self.trial_type_counts['blocks'][block_number] = {
                game_name: 1
            }
        else:
            if game_name not in self.trial_type_counts['blocks'][block_number]:
                self.trial_type_counts['blocks'][block_number][game_name] = 1
            else:
                self.trial_type_counts['blocks'][block_number][game_name] += 1

        if 'itemType' in trial:

            # If itemType is HEALTHY, count them.
            # Also count if this was selected randomly or not
            if trial['itemType'] == 'HEALTHY':
                if 'HEALTHY' not in self.trial_type_counts['blocks'][block_number]:
                    self.trial_type_counts['blocks'][block_number]['HEALTHY'] = 1
                else:
                    self.trial_type_counts['blocks'][block_number]['HEALTHY'] += 1

                if trial['selected'] == 'RANDOM':
                    if 'HEALTHY_RANDOM' not in self.trial_type_counts['blocks'][block_number]:
                        self.trial_type_counts['blocks'][block_number]['HEALTHY_RANDOM'] = 1
                    else:
                        self.trial_type_counts['blocks'][block_number]['HEALTHY_RANDOM'] += 1
                else:
                    if 'HEALTHY_NOT_RANDOM' not in self.trial_type_counts['blocks'][block_number]:
                        self.trial_type_counts['blocks'][block_number]['HEALTHY_NOT_RANDOM'] = 1
                    else:
                        self.trial_type_counts['blocks'][block_number]['HEALTHY_NOT_RANDOM'] += 1

            # If itemType is not healthy, count this
            if trial['itemType'] == 'NON_HEALTHY':
                if 'NON_HEALTHY' not in self.trial_type_counts['blocks'][block_number]:
                    self.trial_type_counts['blocks'][block_number]['NON_HEALTHY'] = 1
                else:
                    self.trial_type_counts['blocks'][block_number]['NON_HEALTHY'] += 1

    def output_global_stats(self):
        print('Called output globals.xlsx')
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Data Output'

        if len(self.session_durations):
            print(self.session_durations)

            sheet.cell(column=1, row=1, value='Min Session Duration')
            sheet.cell(column=2, row=1, value=min(self.session_durations))

            sheet.cell(column=1, row=2, value='Max Session Duration')
            sheet.cell(column=2, row=2, value=max(self.session_durations))

            sheet.cell(column=1, row=3, value='Mean Session Duration')
            sheet.cell(column=2, row=3, value=statistics.mean(self.session_durations))

            sheet.cell(column=1, row=4, value='Session Duration Std Dev')
            sheet.cell(column=2, row=4, value=statistics.stdev(self.session_durations))
        wb.save('globals.xlsx')


if __name__ == '__main__':
    rs = ReadStop()

    # json_file = '/home/ianh/Downloads/STOP.json'
    # rs.read_stop_file(json_file)

    data_dir = '../../'
    abspath = os.path.abspath(data_dir)
    # all_data_file = '200318_new.json'
    # rs.sort_all_data(os.path.join(data_dir, all_data_file))

    all_data_file = '020518.json'
    rs.sort_all_data(os.path.join(data_dir, all_data_file))
    rs.output_global_stats()
